"""
XLSX import for Phase 3B. Deliberately thin: this module's only job is
getting from workbook bytes to (headers, raw_rows) -- the exact same shape
parse_csv_bytes() produces for CSV. Everything downstream (column mapping,
normalization, validation, duplicate detection, summarization) is
app.services.csv_import.process_rows(), unchanged and unduplicated. This
is what guarantees CSV and XLSX behave identically: there is only one
mapping/validation implementation for both to share.

Uses openpyxl (already in requirements.txt from Phase 1) and is fully
testable in this sandbox, since openpyxl has no server/network dependency.
"""

import io
from typing import Optional

import openpyxl

from app.services.csv_import import process_rows


class XlsxParseError(ValueError):
    pass


def list_sheet_names(content: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _detect_header_row(rows: list[tuple]) -> int:
    """
    Returns the 0-based index of the first row with at least 2 non-empty
    cells. Real exports commonly have a title row (1 cell) and/or a blank
    row before the real header -- this skips those without assuming a
    fixed row number. Raises XlsxParseError if no such row exists at all
    within the first 20 rows scanned.
    """
    for i, row in enumerate(rows[:20]):
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
        if non_empty >= 2:
            return i
    raise XlsxParseError("Could not detect a header row (no row in the first 20 has 2+ non-empty cells)")


def parse_xlsx_bytes(content: bytes, sheet_name: Optional[str] = None) -> tuple[str, list[str], list[dict]]:
    """
    Returns (selected_sheet_name, headers, raw_rows). If sheet_name is
    None, uses the workbook's first sheet. Detects the header row rather
    than assuming row 1, so a real export with a title row above its
    header still parses correctly.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        available = list(wb.sheetnames)
        if not available:
            raise XlsxParseError("Workbook has no sheets")
        selected = sheet_name if sheet_name is not None else available[0]
        if selected not in available:
            raise XlsxParseError(f"Sheet '{selected}' not found. Available sheets: {available}")

        ws = wb[selected]
        all_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        if not all_rows:
            raise XlsxParseError(f"Sheet '{selected}' is empty")

        header_idx = _detect_header_row(all_rows)
        header_row = all_rows[header_idx]
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        # Trailing empty header cells (common when a sheet has extra
        # formatting-only columns) are dropped rather than kept as ''.
        while headers and headers[-1] == "":
            headers.pop()

        raw_rows = []
        for row in all_rows[header_idx + 1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue  # skip fully blank rows (common trailing/spacer rows)
            row_dict = {}
            for i, h in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Normalize every cell to a string, same representation
                # parse_csv_bytes() produces via csv.DictReader -- this is
                # what lets process_rows() treat both formats identically
                # (validate_record expects string inputs it parses itself).
                row_dict[h] = "" if value is None else str(value)
            raw_rows.append(row_dict)

        return selected, headers, raw_rows
    finally:
        wb.close()


def run_pipeline(
    content: bytes,
    sheet_name: Optional[str] = None,
    column_mapping: Optional[dict] = None,
) -> dict:
    """
    XLSX entry point, mirroring csv_import.run_pipeline()'s signature and
    return shape exactly, plus the extra sheet_name in/out for workbook
    selection. Delegates all mapping/validation/dedup logic to
    csv_import.process_rows() -- no separate implementation exists.
    """
    selected_sheet, headers, raw_rows = parse_xlsx_bytes(content, sheet_name=sheet_name)
    report = process_rows(headers, raw_rows, column_mapping)
    report["sheet_name"] = selected_sheet
    report["available_sheets"] = list_sheet_names(content)
    return report
