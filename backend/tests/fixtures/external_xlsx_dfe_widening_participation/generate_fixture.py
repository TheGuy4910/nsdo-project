"""
Builds workbook_source.xlsx from the exact, real values transcribed from
the two official DfE dataset preview pages documented in SOURCE.md.

This script is fixture-construction tooling, not part of the application.
Run once to (re)generate the fixture: python3 generate_fixture.py
"""

import openpyxl
import os

wb = openpyxl.Workbook()

# ---- Sheet 1: FSM_Sex_Ethnicity (real title row above the real header, to
# test header-row detection against a non-row-1 header) ----
ws1 = wb.active
ws1.title = "FSM_Sex_Ethnicity"
ws1.append(["Free School Meals, Sex and Ethnic Group -- DfE Widening "
            "participation in higher education, extracted 2026-08-24"])
ws1.append([])  # a genuinely blank row, also real-world-typical before a header
ws1.append([
    "time_period", "time_identifier", "geographic_level", "country_code", "country_name",
    "entry_age", "sex", "fsm_status", "ethnicity_major", "ethnicity_minor",
    "participation_rate", "high_tariff_participation_rate",
    "number_of_he_students", "number_of_high_tariff_he_students", "number_of_students",
])
ws1.append([200506, "Academic year", "National", "E92000001", "England", "By Age 18", "Female",
            "Free School Meals", "White", "English / Welsh / Scottish / Northern Irish / British",
            6.8, 1, 1757, 263, 25794])
ws1.append([200506, "Academic year", "National", "E92000001", "England", "By Age 18", "Female",
            "Free School Meals", "White", "Irish", 6.6, 1.5, 13, 3, 197])
ws1.append([200506, "Academic year", "National", "E92000001", "England", "By Age 18", "Female",
            "Free School Meals", "White", "Traveller of Irish Heritage", 0, 0, 0, 0, 33])
ws1.append([200506, "Academic year", "National", "E92000001", "England", "By Age 18", "Female",
            "Free School Meals", "White", "Gypsy / Roma", 0, 0, 0, 0, 71])
ws1.append([200506, "Academic year", "National", "E92000001", "England", "By Age 18", "Female",
            "Free School Meals", "White", "Any other White background", 18.9, 2.7, 183, 26, 970])

# ---- Sheet 2: All_Characteristics (header on row 1, no title row --
# deliberately different shape from Sheet 1) ----
ws2 = wb.create_sheet("All_Characteristics")
ws2.append([
    "time_period", "time_identifier", "geographic_level", "country_code", "country_name",
    "breakdown_topic", "breakdown", "progression_rate", "high_tariff_progression_rate",
    "number_of_he_students", "number_of_high_tariff_he_students", "number_of_students",
])
ws2.append([200910, "Academic year", "National", "E92000001", "England", "FSM Status",
            "Free School Meals", 18.6, 2.9, 14664, 2297, 78802])
ws2.append([200910, "Academic year", "National", "E92000001", "England", "FSM Status",
            "All Other Pupils", 36.2, 10.9, 187028, 56147, 516955])
ws2.append([200910, "Academic year", "National", "E92000001", "England", "SEN Provision",
            "SEN support / SEN without an EHC plan", 11.2, 1.6, 9505, 1391, 84851])
ws2.append([200910, "Academic year", "National", "E92000001", "England", "SEN Provision",
            "Education, health and care plan", 5.5, 0.9, 1348, 224, 24456])
ws2.append([200910, "Academic year", "National", "E92000001", "England", "SEN Provision",
            "No SEN provision", 39.2, 11.7, 190839, 56829, 486450])

out_path = os.path.join(os.path.dirname(__file__), "workbook_source.xlsx")
wb.save(out_path)
print(f"Wrote {out_path}")
print(f"Sheets: {wb.sheetnames}")
